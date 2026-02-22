"""
RetailPulse Analytics - FastAPI Backend Server
Complete retail analytics pipeline with API endpoints
"""
from fastapi import FastAPI, APIRouter, HTTPException, BackgroundTasks, Query, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse
from starlette.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from typing import List, Dict, Any, Optional
from datetime import datetime, timezone
from pathlib import Path
import os
import logging
import json
import asyncio
import shutil
import pandas as pd
from concurrent.futures import ThreadPoolExecutor
from dotenv import load_dotenv

# Import analytics modules
from analytics.pipeline_orchestrator import PipelineOrchestrator
from analytics.data_ingestion import DataIngestion
from analytics.data_cleaning import DataCleaning
from analytics.feature_engineering import FeatureEngineering
from analytics.eda import ExploratoryDataAnalysis
from analytics.rfm_analysis import RFMAnalysis
from analytics.segmentation import CustomerSegmentation
from analytics.clv import CustomerLifetimeValue
from analytics.kpi_generator import KPIGenerator
from analytics.performance_analysis import PerformanceAnalysis
from analytics.forecasting import SalesForecasting
from analytics.report_generator import ReportGenerator

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Create FastAPI app
app = FastAPI(
    title="RetailPulse Analytics API",
    description="End-to-end retail analytics pipeline",
    version="1.0.0"
)

# API Router with /api prefix
api_router = APIRouter(prefix="/api")

# CORS middleware (FIXED)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://retail-analytics-pipeline.vercel.app",
        "http://localhost:3000"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuration
DATA_DIR = ROOT_DIR / 'data'
OUTPUT_DIR = ROOT_DIR / 'outputs'
UPLOAD_DIR = ROOT_DIR / 'uploads'
DATA_FILE = DATA_DIR / 'Retail_DemoData.csv'

# Required columns for retail analytics
REQUIRED_COLUMNS = [
    'transaction_id', 'quantity', 'transaction_date', 'price',
    'customer_id', 'transaction_amount'
]

OPTIONAL_COLUMNS = [
    'transaction_item_code', 'transaction_item_description',
    'product_type', 'country'
]

# Create directories
DATA_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
UPLOAD_DIR.mkdir(exist_ok=True)

# Thread pool for CPU-intensive tasks
executor = ThreadPoolExecutor(max_workers=4)

# Global state for pipeline
pipeline_state = {
    'status': 'idle',
    'progress': 0,
    'current_stage': None,
    'stages_completed': [],
    'start_time': None,
    'end_time': None,
    'results': None,
    'error': None
}

# Cached results
cached_results = {
    'kpis': None,
    'eda': None,
    'rfm': None,
    'segmentation': None,
    'clv': None,
    'performance': None,
    'forecast': None,
    'last_updated': None
}

# Pydantic Models
class PipelineStatus(BaseModel):
    status: str
    progress: float
    current_stage: Optional[str]
    stages_completed: List[str]
    start_time: Optional[str]
    end_time: Optional[str]
    error: Optional[str]

class PipelineRunRequest(BaseModel):
    stages: Optional[List[str]] = None
    
class ExportRequest(BaseModel):
    format: str = "csv"  # csv or excel
    data_type: str = "cleaned"  # cleaned, rfm, segments, clv


# Helper function to run pipeline
def run_pipeline_sync():
    """Run the analytics pipeline synchronously"""
    global pipeline_state, cached_results
    
    try:
        pipeline_state['status'] = 'running'
        pipeline_state['start_time'] = datetime.now(timezone.utc).isoformat()
        pipeline_state['error'] = None
        
        def progress_callback(update):
            pipeline_state['current_stage'] = update['stage']
            pipeline_state['stages_completed'] = update['stages_completed']
            pipeline_state['progress'] = update['progress']
        
        orchestrator = PipelineOrchestrator(str(DATA_FILE), str(OUTPUT_DIR))
        orchestrator.set_progress_callback(progress_callback)
        
        results = orchestrator.run_full_pipeline()
        
        pipeline_state['status'] = 'completed'
        pipeline_state['end_time'] = datetime.now(timezone.utc).isoformat()
        pipeline_state['progress'] = 100
        pipeline_state['results'] = results
        
        # Cache results
        if 'results' in results:
            cached_results['kpis'] = results['results'].get('kpis')
            cached_results['eda'] = results['results'].get('eda')
            cached_results['rfm'] = results['results'].get('rfm')
            cached_results['segmentation'] = results['results'].get('segmentation')
            cached_results['clv'] = results['results'].get('clv')
            cached_results['performance'] = results['results'].get('performance')
            cached_results['forecast'] = results['results'].get('forecast')
            cached_results['last_updated'] = datetime.now(timezone.utc).isoformat()
        
        return results
        
    except Exception as e:
        logger.error(f"Pipeline error: {str(e)}")
        pipeline_state['status'] = 'failed'
        pipeline_state['error'] = str(e)
        pipeline_state['end_time'] = datetime.now(timezone.utc).isoformat()
        raise


# API Endpoints

@api_router.get("/")
async def root():
    """API root endpoint"""
    return {
        "message": "RetailPulse Analytics API",
        "version": "1.0.0",
        "status": "operational"
    }


@api_router.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "data_file_exists": DATA_FILE.exists()
    }


# Dataset Upload Endpoints

@api_router.post("/data/upload")
async def upload_dataset(file: UploadFile = File(...)):
    """Upload a new dataset file"""
    global DATA_FILE, cached_results
    
    # Validate file extension
    allowed_extensions = ['.xlsx', '.xls', '.csv']
    file_ext = Path(file.filename).suffix.lower()
    
    if file_ext not in allowed_extensions:
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid file type. Allowed: {', '.join(allowed_extensions)}"
        )
    
    # Save uploaded file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_filename = f"dataset_{timestamp}{file_ext}"
    upload_path = UPLOAD_DIR / safe_filename
    
    try:
        with open(upload_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Validate the uploaded file
        validation = await validate_uploaded_file(upload_path)
        
        if not validation['is_valid']:
            # Remove invalid file
            upload_path.unlink(missing_ok=True)
            raise HTTPException(
                status_code=400,
                detail=f"Invalid dataset: {validation['error']}"
            )
        
        # Move to data directory as active dataset
        new_data_file = DATA_DIR / safe_filename
        shutil.move(str(upload_path), str(new_data_file))
        
        # Update global DATA_FILE reference
        DATA_FILE = new_data_file
        
        # Clear cached results
        cached_results = {
            'kpis': None, 'eda': None, 'rfm': None,
            'segmentation': None, 'clv': None,
            'performance': None, 'forecast': None,
            'last_updated': None
        }
        
        return {
            "message": "Dataset uploaded successfully",
            "filename": safe_filename,
            "validation": validation,
            "ready_for_pipeline": True
        }
        
    except HTTPException:
        raise
    except Exception as e:
        upload_path.unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


async def validate_uploaded_file(file_path: Path) -> Dict[str, Any]:
    """Validate uploaded dataset"""
    import pandas as pd
    
    try:
        # Load file
        if file_path.suffix.lower() in ['.xlsx', '.xls']:
            df = pd.read_excel(file_path)
        else:
            df = pd.read_csv(file_path)
        
        validation = {
            'is_valid': True,
            'rows': len(df),
            'columns': len(df.columns),
            'column_names': df.columns.tolist(),
            'missing_required': [],
            'warnings': [],
            'error': None
        }
        
        # Check minimum rows
        if len(df) < 100:
            validation['is_valid'] = False
            validation['error'] = "Dataset too small (minimum 100 rows required)"
            return validation
        
        # Check required columns
        df_columns_lower = [col.lower() for col in df.columns]
        for col in REQUIRED_COLUMNS:
            if col.lower() not in df_columns_lower:
                validation['missing_required'].append(col)
        
        if validation['missing_required']:
            validation['is_valid'] = False
            validation['error'] = f"Missing required columns: {', '.join(validation['missing_required'])}"
            return validation
        
        # Check for data quality issues (warnings only)
        null_pct = (df.isnull().sum() / len(df) * 100).max()
        if null_pct > 50:
            validation['warnings'].append(f"High null percentage detected: {null_pct:.1f}%")
        
        return validation
        
    except Exception as e:
        return {
            'is_valid': False,
            'error': f"Failed to read file: {str(e)}",
            'rows': 0,
            'columns': 0,
            'column_names': [],
            'missing_required': [],
            'warnings': []
        }


@api_router.get("/data/validate")
async def validate_current_dataset():
    """Validate the current active dataset"""
    if not DATA_FILE.exists():
        raise HTTPException(status_code=404, detail="No dataset loaded")
    
    validation = await validate_uploaded_file(DATA_FILE)
    return {
        "filename": DATA_FILE.name,
        "path": str(DATA_FILE),
        "validation": validation
    }


@api_router.get("/data/datasets")
async def list_datasets():
    """List all available datasets"""
    datasets = []
    
    # Check data directory
    for file_path in DATA_DIR.glob('*'):
        if file_path.suffix.lower() in ['.xlsx', '.xls', '.csv']:
            datasets.append({
                'filename': file_path.name,
                'size_mb': round(file_path.stat().st_size / (1024 * 1024), 2),
                'created': datetime.fromtimestamp(file_path.stat().st_mtime).isoformat(),
                'is_active': file_path == DATA_FILE
            })
    
    return sorted(datasets, key=lambda x: x['created'], reverse=True)


@api_router.post("/data/activate/{filename}")
async def activate_dataset(filename: str):
    """Set a dataset as the active dataset for pipeline"""
    global DATA_FILE, cached_results
    
    file_path = DATA_DIR / filename
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Dataset not found")
    
    # Validate before activating
    validation = await validate_uploaded_file(file_path)
    
    if not validation['is_valid']:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot activate invalid dataset: {validation['error']}"
        )
    
    DATA_FILE = file_path
    
    # Clear cached results
    cached_results = {
        'kpis': None, 'eda': None, 'rfm': None,
        'segmentation': None, 'clv': None,
        'performance': None, 'forecast': None,
        'last_updated': None
    }
    
    return {
        "message": f"Dataset '{filename}' activated",
        "validation": validation
    }


# Pipeline Endpoints

@api_router.post("/pipeline/run")
async def run_pipeline(background_tasks: BackgroundTasks):
    """Start the analytics pipeline"""
    global pipeline_state
    
    if pipeline_state['status'] == 'running':
        raise HTTPException(status_code=400, detail="Pipeline is already running")
    
    if not DATA_FILE.exists():
        raise HTTPException(status_code=404, detail="Data file not found")
    
    # Reset state
    pipeline_state = {
        'status': 'starting',
        'progress': 0,
        'current_stage': None,
        'stages_completed': [],
        'start_time': None,
        'end_time': None,
        'results': None,
        'error': None
    }
    
    # Run in background
    background_tasks.add_task(run_pipeline_sync)
    
    return {"message": "Pipeline started", "status": "starting"}


@api_router.get("/pipeline/status", response_model=PipelineStatus)
async def get_pipeline_status():
    """Get current pipeline status"""
    return PipelineStatus(**pipeline_state)


@api_router.get("/pipeline/results")
async def get_pipeline_results():
    """Get pipeline results"""
    if pipeline_state['results'] is None:
        if cached_results['last_updated']:
            return {"cached": True, "results": cached_results}
        raise HTTPException(status_code=404, detail="No results available. Run the pipeline first.")
    
    return pipeline_state['results']


# KPI Endpoints

@api_router.get("/kpis")
async def get_kpis():
    """Get all KPIs"""
    if cached_results['kpis']:
        return cached_results['kpis']
    
    if not DATA_FILE.exists():
        raise HTTPException(status_code=404, detail="Data file not found")
    
    # Generate on-demand
    try:
        ingestion = DataIngestion(str(DATA_FILE))
        df = ingestion.load_data()
        cleaner = DataCleaning(df)
        cleaned = cleaner.run_full_cleaning_pipeline()
        
        kpi = KPIGenerator(cleaned)
        result = kpi.get_kpi_report()
        cached_results['kpis'] = result
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/kpis/summary")
async def get_kpi_summary():
    """Get KPI summary for dashboard"""
    kpis = await get_kpis()
    return kpis.get('summary', [])


# EDA Endpoints

@api_router.get("/eda")
async def get_eda():
    """Get exploratory data analysis"""
    if cached_results['eda']:
        return cached_results['eda']
    
    if not DATA_FILE.exists():
        raise HTTPException(status_code=404, detail="Data file not found")
    
    try:
        ingestion = DataIngestion(str(DATA_FILE))
        df = ingestion.load_data()
        cleaner = DataCleaning(df)
        cleaned = cleaner.run_full_cleaning_pipeline()
        
        eda = ExploratoryDataAnalysis(cleaned)
        result = eda.get_full_eda_report()
        cached_results['eda'] = result
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/eda/top-products")
async def get_top_products(limit: int = Query(10, ge=1, le=50)):
    """Get top products by revenue"""
    eda = await get_eda()
    return eda.get('top_products', [])[:limit]


@api_router.get("/eda/top-customers")
async def get_top_customers(limit: int = Query(10, ge=1, le=50)):
    """Get top customers by spend"""
    eda = await get_eda()
    return eda.get('top_customers', [])[:limit]


# RFM Endpoints

@api_router.get("/rfm")
async def get_rfm():
    """Get RFM analysis"""
    if cached_results['rfm']:
        return cached_results['rfm']
    
    if not DATA_FILE.exists():
        raise HTTPException(status_code=404, detail="Data file not found")
    
    try:
        ingestion = DataIngestion(str(DATA_FILE))
        df = ingestion.load_data()
        cleaner = DataCleaning(df)
        cleaned = cleaner.run_full_cleaning_pipeline()
        
        rfm = RFMAnalysis(cleaned)
        rfm.calculate_rfm_metrics()
        rfm.assign_rfm_scores()
        rfm.assign_customer_segments()
        
        result = rfm.get_rfm_report()
        cached_results['rfm'] = result
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/rfm/segments")
async def get_rfm_segments():
    """Get RFM segment summary"""
    rfm = await get_rfm()
    return rfm.get('segment_summary', [])


@api_router.get("/rfm/distribution")
async def get_rfm_distribution():
    """Get RFM score distributions"""
    rfm = await get_rfm()
    return rfm.get('rfm_distribution', {})


# Segmentation Endpoints

@api_router.get("/segmentation")
async def get_segmentation():
    """Get customer segmentation"""
    if cached_results['segmentation']:
        return cached_results['segmentation']
    
    if not DATA_FILE.exists():
        raise HTTPException(status_code=404, detail="Data file not found")
    
    try:
        ingestion = DataIngestion(str(DATA_FILE))
        df = ingestion.load_data()
        cleaner = DataCleaning(df)
        cleaned = cleaner.run_full_cleaning_pipeline()
        
        seg = CustomerSegmentation(cleaned)
        seg.prepare_customer_data()
        seg.scale_features()
        seg.find_optimal_k()
        seg.perform_clustering()
        
        result = seg.get_segmentation_report()
        cached_results['segmentation'] = result
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/segmentation/clusters")
async def get_cluster_profiles():
    """Get cluster profiles"""
    seg = await get_segmentation()
    return seg.get('cluster_profiles', [])


@api_router.get("/segmentation/elbow")
async def get_elbow_analysis():
    """Get elbow analysis for optimal K"""
    seg = await get_segmentation()
    return seg.get('elbow_analysis', {})


# CLV Endpoints

@api_router.get("/clv")
async def get_clv():
    """Get customer lifetime value analysis"""
    if cached_results['clv']:
        return cached_results['clv']
    
    if not DATA_FILE.exists():
        raise HTTPException(status_code=404, detail="Data file not found")
    
    try:
        ingestion = DataIngestion(str(DATA_FILE))
        df = ingestion.load_data()
        cleaner = DataCleaning(df)
        cleaned = cleaner.run_full_cleaning_pipeline()
        
        clv = CustomerLifetimeValue(cleaned)
        clv.calculate_clv_metrics()
        
        result = clv.get_clv_report()
        cached_results['clv'] = result
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/clv/summary")
async def get_clv_summary():
    """Get CLV summary"""
    clv = await get_clv()
    return clv.get('summary', {})


@api_router.get("/clv/top-customers")
async def get_top_clv_customers(limit: int = Query(10, ge=1, le=50)):
    """Get top customers by CLV"""
    clv = await get_clv()
    return clv.get('top_customers', [])[:limit]


@api_router.get("/clv/at-risk")
async def get_at_risk_customers():
    """Get high-value customers at risk of churning"""
    clv = await get_clv()
    return clv.get('at_risk_customers', [])


# Performance Endpoints

@api_router.get("/performance")
async def get_performance():
    """Get performance analysis"""
    if cached_results['performance']:
        return cached_results['performance']
    
    if not DATA_FILE.exists():
        raise HTTPException(status_code=404, detail="Data file not found")
    
    try:
        ingestion = DataIngestion(str(DATA_FILE))
        df = ingestion.load_data()
        cleaner = DataCleaning(df)
        cleaned = cleaner.run_full_cleaning_pipeline()
        
        perf = PerformanceAnalysis(cleaned)
        result = perf.get_performance_report()
        cached_results['performance'] = result
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/performance/monthly")
async def get_monthly_performance():
    """Get monthly performance"""
    perf = await get_performance()
    return perf.get('monthly', [])


@api_router.get("/performance/category")
async def get_category_performance():
    """Get category performance"""
    perf = await get_performance()
    return perf.get('by_category', [])


@api_router.get("/performance/country")
async def get_country_performance():
    """Get country performance"""
    perf = await get_performance()
    return perf.get('by_country', [])


@api_router.get("/performance/products")
async def get_product_performance(limit: int = Query(20, ge=1, le=100)):
    """Get top products performance"""
    perf = await get_performance()
    return perf.get('top_products', [])[:limit]


# Forecast Endpoints

@api_router.get("/forecast")
async def get_forecast():
    """Get sales forecast"""
    if cached_results['forecast']:
        return cached_results['forecast']
    
    if not DATA_FILE.exists():
        raise HTTPException(status_code=404, detail="Data file not found")
    
    try:
        ingestion = DataIngestion(str(DATA_FILE))
        df = ingestion.load_data()
        cleaner = DataCleaning(df)
        cleaned = cleaner.run_full_cleaning_pipeline()
        
        forecast = SalesForecasting(cleaned)
        result = forecast.get_forecast_report(forecast_periods=6)
        cached_results['forecast'] = result
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/forecast/comparison")
async def get_forecast_comparison():
    """Get forecast method comparison"""
    forecast = await get_forecast()
    return forecast.get('comparison', {})


@api_router.get("/forecast/seasonal")
async def get_seasonal_analysis():
    """Get seasonal analysis"""
    forecast = await get_forecast()
    return forecast.get('seasonal_analysis', {})


# Export Endpoints

@api_router.get("/exports")
async def list_exports():
    """List available export files"""
    exports = []
    
    if OUTPUT_DIR.exists():
        for file_path in OUTPUT_DIR.glob('*'):
            if file_path.is_file():
                exports.append({
                    'filename': file_path.name,
                    'size_kb': round(file_path.stat().st_size / 1024, 2),
                    'created': datetime.fromtimestamp(file_path.stat().st_mtime).isoformat(),
                    'type': file_path.suffix[1:].upper() if file_path.suffix else 'UNKNOWN'
                })
    
    return sorted(exports, key=lambda x: x['created'], reverse=True)


@api_router.get("/exports/{filename}")
async def download_export(filename: str):
    """Download an export file"""
    file_path = OUTPUT_DIR / filename
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        path=str(file_path),
        filename=filename,
        media_type='application/octet-stream'
    )


@api_router.post("/exports/generate")
async def generate_export(request: ExportRequest, background_tasks: BackgroundTasks):
    """Generate a new export file"""
    if not DATA_FILE.exists():
        raise HTTPException(status_code=404, detail="Data file not found")
    
    try:
        # Load and process data
        ingestion = DataIngestion(str(DATA_FILE))
        df = ingestion.load_data()
        cleaner = DataCleaning(df)
        cleaned = cleaner.run_full_cleaning_pipeline()
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if request.data_type == 'cleaned':
            data = cleaned
            base_name = f'cleaned_data_{timestamp}'
        elif request.data_type == 'rfm':
            rfm = RFMAnalysis(cleaned)
            rfm.calculate_rfm_metrics()
            rfm.assign_rfm_scores()
            rfm.assign_customer_segments()
            data = rfm.get_rfm_data()
            base_name = f'rfm_analysis_{timestamp}'
        elif request.data_type == 'segments':
            seg = CustomerSegmentation(cleaned)
            seg.prepare_customer_data()
            seg.scale_features()
            seg.find_optimal_k()
            seg.perform_clustering()
            data = seg.get_segmented_customers()
            base_name = f'customer_segments_{timestamp}'
        elif request.data_type == 'clv':
            clv = CustomerLifetimeValue(cleaned)
            clv.calculate_clv_metrics()
            data = clv.get_clv_data()
            base_name = f'customer_ltv_{timestamp}'
        else:
            raise HTTPException(status_code=400, detail="Invalid data_type")
        
        # Export
        if request.format == 'csv':
            filepath = OUTPUT_DIR / f'{base_name}.csv'
            data.to_csv(filepath, index=False)
        elif request.format == 'excel':
            filepath = OUTPUT_DIR / f'{base_name}.xlsx'
            data.to_excel(filepath, index=False)
        else:
            raise HTTPException(status_code=400, detail="Invalid format")
        
        return {
            "message": "Export generated successfully",
            "filename": filepath.name,
            "path": str(filepath)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# Report Endpoints

@api_router.post("/reports/generate")
async def generate_report():
    """Generate an HTML report"""
    if not DATA_FILE.exists():
        raise HTTPException(status_code=404, detail="Data file not found")
    
    try:
        # Compile all analytics data
        report_data = {
            'kpis': (await get_kpis()).get('kpis', {}),
            'summary': (await get_kpis()).get('summary', []),
            'rfm': await get_rfm(),
            'segmentation': await get_segmentation(),
            'performance': await get_performance(),
            'forecast': await get_forecast()
        }
        
        reporter = ReportGenerator(str(OUTPUT_DIR))
        html_path = reporter.generate_html_report(report_data, "Retail Analytics Report")
        
        return {
            "message": "Report generated successfully",
            "filename": Path(html_path).name,
            "path": html_path
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/reports")
async def list_reports():
    """List generated reports"""
    reports = []
    
    if OUTPUT_DIR.exists():
        for file_path in OUTPUT_DIR.glob('report_*.html'):
            reports.append({
                'filename': file_path.name,
                'size_kb': round(file_path.stat().st_size / 1024, 2),
                'created': datetime.fromtimestamp(file_path.stat().st_mtime).isoformat()
            })
    
    return sorted(reports, key=lambda x: x['created'], reverse=True)


# Cached data info
cached_data_info = {
    'info': None,
    'file_path': None
}

# Data Info Endpoint

@api_router.get("/data/info")
async def get_data_info():
    """Get information about the loaded data"""
    global cached_data_info
    
    if not DATA_FILE.exists():
        return {
            "exists": False,
            "message": "Data file not found"
        }
    
    # Return cached if same file
    if cached_data_info['file_path'] == str(DATA_FILE) and cached_data_info['info']:
        return cached_data_info['info']
    
    try:
        import pandas as pd
        
        # Quick row count without loading entire file
        if DATA_FILE.suffix.lower() in ['.xlsx', '.xls']:
            # For Excel, load but don't process
            df = pd.read_excel(DATA_FILE, nrows=0)  # Just headers
            # Use openpyxl to get row count efficiently
            from openpyxl import load_workbook
            wb = load_workbook(DATA_FILE, read_only=True, data_only=True)
            ws = wb.active
            row_count = ws.max_row - 1  # Subtract header
            wb.close()
            columns = df.columns.tolist()
        else:
            # For CSV, count lines
            with open(DATA_FILE, 'r') as f:
                row_count = sum(1 for _ in f) - 1  # Subtract header
            df = pd.read_csv(DATA_FILE, nrows=0)
            columns = df.columns.tolist()
        
        result = {
            "exists": True,
            "filename": DATA_FILE.name,
            "size_mb": round(DATA_FILE.stat().st_size / (1024 * 1024), 2),
            "rows": row_count,
            "columns": len(columns),
            "column_names": columns,
            "validation": {
                "is_valid": True,
                "total_records": row_count,
                "total_columns": len(columns)
            }
        }
        
        # Cache result
        cached_data_info['info'] = result
        cached_data_info['file_path'] = str(DATA_FILE)
        
        return result
        
    except Exception as e:
        logger.error(f"Error getting data info: {str(e)}")
        # Return basic info on error
        return {
            "exists": True,
            "filename": DATA_FILE.name,
            "size_mb": round(DATA_FILE.stat().st_size / (1024 * 1024), 2),
            "rows": 0,
            "columns": 0,
            "column_names": [],
            "error": str(e)
        }


# Include router
app.include_router(api_router)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
